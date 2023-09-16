import openpyxl

# Open the source Excel file for reading
source_workbook = openpyxl.load_workbook('source_file.xlsx')
source_sheet = source_workbook.active  # Assuming you want to work with the active sheet

# Create a new Excel file for writing
target_workbook = openpyxl.Workbook()
target_sheet = target_workbook.active

# Read data from the source sheet and write it to the target sheet
for row in source_sheet.iter_rows(values_only=True):
    target_sheet.append(row)  # Append each row to the target sheet

# Save the target Excel file
target_workbook.save('target_file.xlsx')

# Close both workbooks
source_workbook.close()
target_workbook.close()

print("Data has been copied from source_file.xlsx to target_file.xlsx")